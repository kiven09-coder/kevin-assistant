"""
🤖 Kiven — Arabic Personal AI Assistant
========================================
A Jarvis-style multimodal assistant for Amr.

Capabilities:
- 🎙️ Arabic voice (Whisper STT + Edge-TTS, Egyptian dialect)
- 🧠 Dual LLM brain: Claude (smart, tool-use) + Gemini (fast, free fallback)
- 🔧 Tools: web_search, generate_image, remember/recall, time, weather, calculate
- 📚 Excel-backed knowledge base
- 🔒 Password-gated (HMAC-signed sessions)
- 💾 Persistent long-term memory across sessions

Architecture mirrors Microsoft HuggingGPT:
  LLM controller → TOOLS_REGISTRY → multimodal output
"""

import os
import sys
import asyncio
import tempfile
import json
import socket
import glob
import hmac
import hashlib
import secrets
import base64
from datetime import datetime
from pathlib import Path

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
        sys.stderr.reconfigure(encoding="utf-8")
    except Exception:
        pass

# ============================================
# Load .env file
# ============================================
def load_env_file():
    env_path = Path(".env")
    if env_path.exists():
        for line in env_path.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            if "=" in line:
                key, value = line.split("=", 1)
                key = key.strip()
                value = value.strip().strip('"').strip("'")
                if key and not os.getenv(key):
                    os.environ[key] = value
        print("✅ تم تحميل ملف .env")

load_env_file()

from fastapi import FastAPI, File, UploadFile, HTTPException, Request, Response, Cookie, Depends
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse, RedirectResponse
from fastapi.middleware.cors import CORSMiddleware
import uvicorn

from anthropic import Anthropic
import google.generativeai as genai

# ============================================
# Resolve FFmpeg binary (Whisper requires it on PATH)
# ============================================
def _ensure_imageio_ffmpeg():
    """Import imageio_ffmpeg; if missing, auto-install via pip and retry once."""
    try:
        import imageio_ffmpeg
        return imageio_ffmpeg
    except ImportError:
        pass
    print("⏬ imageio-ffmpeg مش موجودة - جاري تنصيبها أوتوماتيك (مرة واحدة)...")
    import subprocess
    try:
        subprocess.check_call(
            [sys.executable, "-m", "pip", "install", "--quiet", "imageio-ffmpeg>=0.5.0"],
            timeout=300,
        )
    except Exception as e:
        print(f"❌ فشل تنصيب imageio-ffmpeg: {e}")
        return None
    try:
        import importlib
        import imageio_ffmpeg
        importlib.invalidate_caches()
        print("✅ imageio-ffmpeg اتنصبت بنجاح")
        return imageio_ffmpeg
    except ImportError as e:
        print(f"❌ لسة مش قادر يـ import بعد التنصيب: {e}")
        return None

def _setup_ffmpeg():
    try:
        import shutil
        imageio_ffmpeg = _ensure_imageio_ffmpeg()
        if imageio_ffmpeg is None:
            print("⚠️ هـ Whisper مش هيشتغل لحد ما FFmpeg يكون متوفر. شغل يدوياً: pip install imageio-ffmpeg")
            return
        src = imageio_ffmpeg.get_ffmpeg_exe()
        if not src or not os.path.exists(src):
            print(f"⚠️ ffmpeg binary مش موجود في: {src}")
            return
        cache = Path.home() / ".kevin_cache"
        cache.mkdir(exist_ok=True)
        target = cache / ("ffmpeg.exe" if sys.platform == "win32" else "ffmpeg")
        src_size = Path(src).stat().st_size
        if not target.exists() or target.stat().st_size != src_size:
            shutil.copy(src, target)
            if sys.platform != "win32":
                os.chmod(target, 0o755)
        os.environ["PATH"] = str(cache) + os.pathsep + os.environ.get("PATH", "")
        # Verify ffmpeg is actually callable now (definitive check)
        try:
            import subprocess
            subprocess.run(
                [str(target), "-version"],
                capture_output=True, timeout=5, check=True,
            )
            print(f"✅ FFmpeg جاهز ومـ verified: {target}")
        except Exception as e:
            print(f"⚠️ FFmpeg متنصب بس مش قادر يشتغل: {e}")
    except Exception as e:
        print(f"⚠️ مشكلة في تجهيز FFmpeg: {type(e).__name__}: {e}")

_setup_ffmpeg()

import whisper
import edge_tts

# ============================================
# CONFIGURATION
# ============================================
ASSISTANT_NAME = "Kiven"
ASSISTANT_NAME_AR = "كيفن"
ASSISTANT_VERSION = "0.3.0"

ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY", "").strip()
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY", "").strip()
KEVIN_PASSWORD = os.getenv("KEVIN_PASSWORD", "").strip()
KEVIN_SESSION_SECRET = os.getenv("KEVIN_SESSION_SECRET", "").strip()

if not KEVIN_PASSWORD:
    print("⚠️ تحذير: KEVIN_PASSWORD غير محدد في .env - Kevin لن يقبل الدخول!")
if not KEVIN_SESSION_SECRET:
    print("⚠️ تحذير: KEVIN_SESSION_SECRET غير محدد - سيتم استخدام مفتاح مؤقت")
    KEVIN_SESSION_SECRET = secrets.token_urlsafe(32)

SESSION_COOKIE_NAME = "kevin_session"
SESSION_TTL_SECONDS = 60 * 60 * 24 * 30  # 30 days

CLAUDE_MODEL = "claude-sonnet-4-5"
GEMINI_MODEL = "gemini-2.0-flash"  # 200 req/day free tier vs 20 for 2.5-flash
WHISPER_MODEL = "medium"

ARABIC_VOICES = {
    "salma_egypt": "ar-EG-SalmaNeural",
    "shakir_egypt": "ar-EG-ShakirNeural",
    "zariyah_saudi": "ar-SA-ZariyahNeural",
    "hamed_saudi": "ar-SA-HamedNeural",
}
DEFAULT_VOICE = "shakir_egypt"  # Male voice for Kevin

HOST = "0.0.0.0"
PORT = 8000

CONVERSATION_FILE = Path("conversation_history.json")
KNOWLEDGE_BASE_FILE = Path("knowledge_base.json")

# ============================================
# KNOWLEDGE BASE - Load Excel files
# ============================================
def load_knowledge_base():
    """Load all Excel files in current directory as knowledge base."""
    kb = {"files": [], "summary": ""}

    # Check if cached version exists
    if KNOWLEDGE_BASE_FILE.exists():
        try:
            cached = json.loads(KNOWLEDGE_BASE_FILE.read_text(encoding="utf-8"))
            print(f"✅ تم تحميل قاعدة المعرفة المخزنة ({len(cached.get('files', []))} ملف)")
            return cached
        except Exception:
            pass

    search_dirs = ["upload/data", "data", "."]
    excel_files = []
    for d in search_dirs:
        excel_files.extend(glob.glob(os.path.join(d, "*.xlsx")))
        excel_files.extend(glob.glob(os.path.join(d, "*.xls")))
    seen = set()
    unique_files = []
    for f in excel_files:
        key = os.path.basename(f).lower()
        if key not in seen:
            seen.add(key)
            unique_files.append(f)
    excel_files = unique_files

    if not excel_files:
        print("ℹ️ مفيش ملفات Excel في الفولدر (شوف upload/data/)")
        return kb

    print(f"📚 جاري قراءة {len(excel_files)} ملف Excel...")

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("⚠️ openpyxl مش متنصبة. شغل: pip install openpyxl")
        return kb

    for excel_file in excel_files:
        try:
            print(f"   📄 قراءة: {excel_file}")
            wb = load_workbook(excel_file, data_only=True)
            file_data = {
                "filename": excel_file,
                "sheets": []
            }

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                for row in ws.iter_rows(values_only=True, max_row=200):  # Limit to 200 rows per sheet
                    if any(cell is not None and str(cell).strip() for cell in row):
                        rows.append([str(c) if c is not None else "" for c in row])

                file_data["sheets"].append({
                    "name": sheet_name,
                    "row_count": len(rows),
                    "rows": rows[:50]  # Keep only first 50 rows for context
                })
            kb["files"].append(file_data)
        except Exception as e:
            print(f"   ⚠️ خطأ في {excel_file}: {e}")

    # Build a COMPACT text summary for AI context (just file/sheet names + column headers)
    # Sample rows are kept in the JSON cache but NOT injected in every LLM call,
    # to stay under per-minute input-token rate limits.
    summary_parts = [f"📚 قاعدة المعرفة المتاحة (من {len(kb['files'])} ملف Excel):\n"]
    for f in kb["files"]:
        summary_parts.append(f"\n📄 {f['filename']}")
        for sheet in f["sheets"]:
            line = f"  • {sheet['name']} ({sheet['row_count']} صف)"
            if sheet["rows"]:
                cols = " | ".join(sheet["rows"][0][:6])
                if cols.strip():
                    line += f" — أعمدة: {cols}"
            summary_parts.append(line)

    kb["summary"] = "\n".join(summary_parts)

    # Cache for next time
    try:
        KNOWLEDGE_BASE_FILE.write_text(
            json.dumps(kb, ensure_ascii=False, indent=2),
            encoding="utf-8"
        )
    except Exception:
        pass

    print(f"✅ تم تحميل {len(kb['files'])} ملف")
    return kb

knowledge_base = load_knowledge_base()

# ============================================
# TOOLS - Web Search (Phase 1) + extensible registry for future connectors
# ============================================
def _format_search_results(results: list, query: str, kind: str) -> str:
    if not results:
        return f"مفيش نتائج لـ '{query}'."
    label = "أخبار" if kind == "news" else "نتائج البحث"
    lines = [f"📰 {label} عن '{query}':\n"]
    for i, r in enumerate(results, 1):
        title = (r.get("title") or "").strip()
        url = (r.get("href") or r.get("url") or "").strip()
        snippet = (r.get("body") or "").strip()[:300]
        date = (r.get("date") or "").strip()
        chunk = f"{i}. {title}"
        if date:
            chunk += f"  ({date})"
        if snippet:
            chunk += f"\n   {snippet}"
        if url:
            chunk += f"\n   🔗 {url}"
        lines.append(chunk)
    return "\n".join(lines)

def _search_sync(query: str, kind: str, max_results: int) -> list:
    from ddgs import DDGS  # raises ImportError if missing
    with DDGS() as ddgs:
        if kind == "news":
            return list(ddgs.news(query, max_results=max_results, safesearch="off"))
        return list(ddgs.text(query, max_results=max_results, safesearch="off"))

async def web_search(query: str, kind: str = "web", max_results: int = 5) -> str:
    """Search the web or news headlines via DuckDuckGo. No API key required."""
    try:
        max_results = max(1, min(int(max_results), 10))
        kind = kind if kind in ("web", "news") else "web"
        print(f"🔍 Web search [{kind}]: '{query}' (max {max_results})")
        results = await asyncio.to_thread(_search_sync, query, kind, max_results)
        return _format_search_results(results, query, kind)
    except ImportError:
        return "خطأ: حزمة ddgs مش متنصبة. شغل: pip install ddgs"
    except Exception as e:
        print(f"❌ Web search error: {type(e).__name__}: {e}")
        return f"خطأ في البحث: {type(e).__name__}: {str(e)[:200]}"

# ============================================
# Persistent Memory (JSON-backed, no DB)
# ============================================
MEMORY_FILE = Path("kiven_memory.json")
_MAX_FACTS = 200

def _load_memory() -> dict:
    if MEMORY_FILE.exists():
        try:
            return json.loads(MEMORY_FILE.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {"facts": [], "next_id": 1}

def _save_memory(mem: dict) -> None:
    MEMORY_FILE.write_text(json.dumps(mem, ensure_ascii=False, indent=2), encoding="utf-8")

def memory_summary_for_prompt() -> str:
    mem = _load_memory()
    facts = mem.get("facts", [])
    if not facts:
        return ""
    lines = ["💾 ذاكرة Kiven عن المستخدم (من جلسات سابقة):"]
    for f in facts[-30:]:  # last 30 facts max in prompt
        lines.append(f"  • [{f['id']}] {f['text']}")
    return "\n".join(lines)

async def remember(fact: str) -> str:
    fact = (fact or "").strip()
    if not fact:
        return "خطأ: مفيش حاجة أحفظها."
    if len(fact) > 500:
        fact = fact[:500] + "..."
    mem = _load_memory()
    fid = mem["next_id"]
    mem["facts"].append({
        "id": fid,
        "text": fact,
        "saved_at": datetime.now().isoformat(timespec="seconds"),
    })
    mem["next_id"] = fid + 1
    # Trim to max
    if len(mem["facts"]) > _MAX_FACTS:
        mem["facts"] = mem["facts"][-_MAX_FACTS:]
    _save_memory(mem)
    print(f"💾 Remembered #{fid}: {fact[:80]}")
    return f"✅ اتحفظ في الذاكرة (رقم {fid}): {fact}"

async def recall(query: str = "", limit: int = 10) -> str:
    mem = _load_memory()
    facts = mem.get("facts", [])
    if not facts:
        return "الذاكرة فاضية لسة."
    query_low = (query or "").lower().strip()
    if query_low:
        matches = [f for f in facts if query_low in f["text"].lower()]
    else:
        matches = facts
    matches = matches[-int(limit or 10):]
    if not matches:
        return f"مفيش حاجة في الذاكرة عن '{query}'."
    lines = ["📚 من الذاكرة:"]
    for f in matches:
        lines.append(f"  • [{f['id']}] {f['text']}  (محفوظ: {f['saved_at'][:10]})")
    return "\n".join(lines)

async def forget(fact_id: int) -> str:
    mem = _load_memory()
    before = len(mem["facts"])
    mem["facts"] = [f for f in mem["facts"] if f["id"] != int(fact_id)]
    if len(mem["facts"]) == before:
        return f"مش لاقي fact بالرقم {fact_id}."
    _save_memory(mem)
    return f"🗑️ اتمسح من الذاكرة الرقم {fact_id}."

# ============================================
# Time / Date
# ============================================
async def get_current_time(timezone: str = "Africa/Cairo") -> str:
    try:
        # Windows needs tzdata package; gracefully fall back to system local time.
        tz = None
        try:
            from zoneinfo import ZoneInfo
            try:
                tz = ZoneInfo(timezone)
            except Exception:
                try:
                    tz = ZoneInfo("Africa/Cairo")
                except Exception:
                    tz = None
        except ImportError:
            tz = None
        now = datetime.now(tz) if tz else datetime.now()
        tz_label = timezone if tz else "ساعة الجهاز المحلية"
        days_ar = ["الإثنين", "الثلاثاء", "الأربعاء", "الخميس", "الجمعة", "السبت", "الأحد"]
        months_ar = ["", "يناير", "فبراير", "مارس", "أبريل", "مايو", "يونيو",
                     "يوليو", "أغسطس", "سبتمبر", "أكتوبر", "نوفمبر", "ديسمبر"]
        day_name = days_ar[now.weekday()]
        month_name = months_ar[now.month]
        return (
            f"📅 {day_name}، {now.day} {month_name} {now.year}\n"
            f"🕐 الساعة {now.strftime('%I:%M %p')} ({tz_label})"
        )
    except Exception as e:
        return f"خطأ في قراءة الوقت: {e}"

# ============================================
# Weather (Open-Meteo, free, no key)
# ============================================
WEATHER_CODES_AR = {
    0: "صافي ☀️", 1: "غايم خفيف 🌤️", 2: "غايم جزئي ⛅", 3: "غايم 🌥️",
    45: "ضباب 🌫️", 48: "ضباب مع صقيع 🌫️",
    51: "رذاذ خفيف 🌦️", 53: "رذاذ متوسط 🌦️", 55: "رذاذ كثيف 🌧️",
    61: "مطر خفيف 🌧️", 63: "مطر متوسط 🌧️", 65: "مطر شديد 🌧️",
    71: "ثلج خفيف 🌨️", 73: "ثلج متوسط 🌨️", 75: "ثلج كثيف 🌨️",
    80: "زخات مطر 🌦️", 81: "زخات مطر 🌧️", 82: "زخات مطر شديدة ⛈️",
    95: "عاصفة رعدية ⛈️", 96: "عاصفة برعد وبَرَد ⛈️", 99: "عاصفة قوية ⛈️",
}

def _weather_sync(city: str) -> str:
    import urllib.request, urllib.parse
    # Geocode
    g_url = f"https://geocoding-api.open-meteo.com/v1/search?name={urllib.parse.quote(city)}&count=1&language=ar"
    with urllib.request.urlopen(g_url, timeout=10) as r:
        g = json.loads(r.read().decode("utf-8"))
    results = g.get("results") or []
    if not results:
        return f"مش لاقي مدينة اسمها '{city}'."
    loc = results[0]
    lat, lon = loc["latitude"], loc["longitude"]
    name = loc.get("name", city)
    country = loc.get("country", "")
    # Forecast
    w_url = (f"https://api.open-meteo.com/v1/forecast?latitude={lat}&longitude={lon}"
             f"&current=temperature_2m,relative_humidity_2m,weather_code,wind_speed_10m"
             f"&daily=temperature_2m_max,temperature_2m_min&timezone=auto")
    with urllib.request.urlopen(w_url, timeout=10) as r:
        w = json.loads(r.read().decode("utf-8"))
    cur = w.get("current") or {}
    daily = w.get("daily") or {}
    code = int(cur.get("weather_code", 0))
    cond = WEATHER_CODES_AR.get(code, f"حالة جوية كود {code}")
    t = cur.get("temperature_2m")
    h = cur.get("relative_humidity_2m")
    wind = cur.get("wind_speed_10m")
    t_max = (daily.get("temperature_2m_max") or [None])[0]
    t_min = (daily.get("temperature_2m_min") or [None])[0]
    return (
        f"🌍 الطقس في {name}{', ' + country if country else ''}:\n"
        f"  {cond}\n"
        f"  🌡️ درجة الحرارة: {t}°C (الصغرى {t_min}°C / الكبرى {t_max}°C اليوم)\n"
        f"  💧 الرطوبة: {h}%\n"
        f"  💨 الرياح: {wind} كم/ساعة"
    )

async def get_weather(city: str) -> str:
    try:
        city = (city or "").strip() or "Cairo"
        print(f"🌤️ Weather: {city}")
        return await asyncio.to_thread(_weather_sync, city)
    except Exception as e:
        print(f"❌ Weather error: {type(e).__name__}: {e}")
        return f"خطأ في جلب الطقس: {type(e).__name__}: {str(e)[:120]}"

# Arabic city name → Open-Meteo geocoding key. Used to skip an LLM call.
ARABIC_CITY_MAP = {
    "القاهرة": "Cairo", "القاهره": "Cairo", "كايرو": "Cairo",
    "الإسكندرية": "Alexandria", "الاسكندرية": "Alexandria",
    "اسكندرية": "Alexandria", "إسكندرية": "Alexandria",
    "الجيزة": "Giza", "جيزة": "Giza", "الجيزه": "Giza",
    "شرم الشيخ": "Sharm El Sheikh", "شرم": "Sharm El Sheikh",
    "الغردقة": "Hurghada", "غردقة": "Hurghada", "الغردقه": "Hurghada",
    "المنصورة": "Mansoura", "منصورة": "Mansoura",
    "طنطا": "Tanta", "الزقازيق": "Zagazig", "بنها": "Banha",
    "الأقصر": "Luxor", "اقصر": "Luxor", "الاقصر": "Luxor",
    "أسوان": "Aswan", "اسوان": "Aswan",
    "بورسعيد": "Port Said", "بور سعيد": "Port Said",
    "السويس": "Suez", "الإسماعيلية": "Ismailia", "إسماعيلية": "Ismailia",
    "أسيوط": "Asyut", "اسيوط": "Asyut",
    "المنيا": "Minya", "بني سويف": "Beni Suef",
    "دبي": "Dubai", "الرياض": "Riyadh", "رياض": "Riyadh",
    "جدة": "Jeddah", "جده": "Jeddah", "مكة": "Mecca", "المدينة": "Medina",
    "بيروت": "Beirut", "عمان": "Amman", "دمشق": "Damascus", "بغداد": "Baghdad",
    "الدوحة": "Doha", "دوحة": "Doha",
    "أبوظبي": "Abu Dhabi", "ابوظبي": "Abu Dhabi", "أبو ظبي": "Abu Dhabi",
    "الكويت": "Kuwait", "المنامة": "Manama", "البحرين": "Manama", "مسقط": "Muscat",
    "تونس": "Tunis", "الجزائر": "Algiers",
    "الرباط": "Rabat", "الدار البيضاء": "Casablanca", "كازابلانكا": "Casablanca",
    "الخرطوم": "Khartoum",
}

def extract_cities_from_message(message: str) -> list[str]:
    """Find all known Arabic city mentions in a message. Falls back to ['Cairo']."""
    found = []
    # Sort by length desc so longer matches win (e.g. "شرم الشيخ" before "شرم")
    for ar in sorted(ARABIC_CITY_MAP.keys(), key=len, reverse=True):
        if ar in message:
            en = ARABIC_CITY_MAP[ar]
            if en not in found:
                found.append(en)
    return found or ["Cairo"]

# ============================================
# Calculator (safe AST eval)
# ============================================
import ast as _ast
import operator as _op
_SAFE_BIN_OPS = {
    _ast.Add: _op.add, _ast.Sub: _op.sub, _ast.Mult: _op.mul,
    _ast.Div: _op.truediv, _ast.FloorDiv: _op.floordiv,
    _ast.Mod: _op.mod, _ast.Pow: _op.pow,
}
_SAFE_UNARY_OPS = {_ast.UAdd: _op.pos, _ast.USub: _op.neg}

def _safe_eval(node):
    if isinstance(node, _ast.Expression):
        return _safe_eval(node.body)
    if isinstance(node, _ast.Constant) and isinstance(node.value, (int, float)):
        return node.value
    if isinstance(node, _ast.BinOp) and type(node.op) in _SAFE_BIN_OPS:
        return _SAFE_BIN_OPS[type(node.op)](_safe_eval(node.left), _safe_eval(node.right))
    if isinstance(node, _ast.UnaryOp) and type(node.op) in _SAFE_UNARY_OPS:
        return _SAFE_UNARY_OPS[type(node.op)](_safe_eval(node.operand))
    raise ValueError(f"unsupported expression: {_ast.dump(node)}")

async def calculate(expression: str) -> str:
    expr = (expression or "").strip()
    if not expr:
        return "خطأ: مفيش معادلة."
    # Allow common Arabic separators
    expr = expr.replace("×", "*").replace("÷", "/").replace("،", ",")
    if len(expr) > 200:
        return "خطأ: المعادلة طويلة جداً."
    try:
        tree = _ast.parse(expr, mode="eval")
        result = _safe_eval(tree)
        if isinstance(result, float):
            # Show as float only if non-integer
            if result.is_integer():
                result = int(result)
            else:
                result = round(result, 10)
        return f"🧮 {expr} = **{result}**"
    except ZeroDivisionError:
        return "خطأ: قسمة على صفر."
    except Exception as e:
        return f"خطأ في المعادلة: {type(e).__name__}: {str(e)[:120]}"

async def generate_image(prompt: str, width: int = 1024, height: int = 1024) -> str:
    """Generate an image via Pollinations.ai. Returns Markdown with embedded image URL."""
    try:
        import urllib.parse
        prompt_clean = (prompt or "").strip()
        if not prompt_clean:
            return "خطأ: لازم تحدد وصف الصورة."
        width = max(256, min(int(width), 2048))
        height = max(256, min(int(height), 2048))
        seed = secrets.randbelow(1_000_000)
        encoded = urllib.parse.quote(prompt_clean[:500])
        url = (
            f"https://image.pollinations.ai/prompt/{encoded}"
            f"?width={width}&height={height}&seed={seed}&nologo=true&model=flux"
        )
        print(f"🎨 Image gen: '{prompt_clean[:80]}' [{width}x{height}, seed={seed}]")
        return f"![{prompt_clean[:120]}]({url})\n\n*🎨 صورة من Pollinations AI (Flux)*"
    except Exception as e:
        print(f"❌ generate_image error: {type(e).__name__}: {e}")
        return f"خطأ في إنشاء الصورة: {type(e).__name__}: {str(e)[:120]}"

TOOLS_REGISTRY = {
    "web_search": {
        "name": "web_search",
        "description": (
            "ابحث في الويب أو الأخبار عبر DuckDuckGo. "
            "استخدم هذه الأداة لما المستخدم يسأل عن: أحداث جارية، آخر الأخبار، "
            "أسعار حالية، طقس، نتائج رياضية، أو أي معلومة محتاجة بيانات حديثة "
            "من النت مش موجودة في تدريبك. "
            "استخدم kind='news' للأخبار الحديثة، kind='web' للبحث العام."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "query": {
                    "type": "string",
                    "description": "نص البحث (عربي أو إنجليزي حسب الموضوع)."
                },
                "kind": {
                    "type": "string",
                    "enum": ["web", "news"],
                    "description": "نوع البحث.",
                },
                "max_results": {
                    "type": "integer",
                    "description": "عدد النتائج (1-10). الافتراضي 5.",
                    "minimum": 1,
                    "maximum": 10,
                },
            },
            "required": ["query"],
        },
        "handler": web_search,
    },
    "generate_image": {
        "name": "generate_image",
        "description": (
            "أنشئ صورة من وصف نصي (text-to-image) عبر Pollinations AI / Flux. "
            "استخدم هذه الأداة لما المستخدم يطلب: صورة، كرت بصورة، رسمة، تصميم، "
            "لوحة، خلفية، أو أي محتوى بصري. "
            "**مهم جداً: الـ prompt يجب أن يكون بالإنجليزية** عشان جودة أعلى. "
            "إذا طلب المستخدم بالعربية، حوّل في عقلك لوصف إنجليزي مفصل: "
            "(الموضوع، الأسلوب، الألوان، الإضاءة، المزاج). "
            "للكروت/البطاقات: استخدم 'A beautiful greeting card with...' "
            "أبعاد افتراضية 1024×1024 (مربع، مثالي للكروت). "
            "بعد التنفيذ، الأداة بتـ return Markdown image link — ضعه في ردك كما هو "
            "وأكتب نص المعايدة أو الوصف بالعربية بجانبه."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "prompt": {
                    "type": "string",
                    "description": "Detailed English image prompt. Be specific about subject, style, colors, mood, lighting.",
                },
                "width": {
                    "type": "integer",
                    "minimum": 256, "maximum": 2048,
                    "description": "Width in pixels. 1024=square (cards). 1536=landscape. 768=portrait.",
                },
                "height": {
                    "type": "integer",
                    "minimum": 256, "maximum": 2048,
                    "description": "Height in pixels. 1024=square (cards). 768=landscape. 1536=portrait.",
                },
            },
            "required": ["prompt"],
        },
        "handler": generate_image,
    },
    "remember": {
        "name": "remember",
        "description": (
            "احفظ معلومة دائمة عن المستخدم (تاريخه، تفضيلاته، أسماء أهله، أهدافه، "
            "أحداث مهمة، أرقام يحب يفتكرها). الذاكرة دي بتفضل بين الجلسات. "
            "استخدمها لما المستخدم يقول 'افتكر إن...' أو لما تتعرف على معلومة شخصية مهمة."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "fact": {"type": "string", "description": "المعلومة بالعربي، جملة واحدة واضحة."},
            },
            "required": ["fact"],
        },
        "handler": remember,
    },
    "recall": {
        "name": "recall",
        "description": (
            "ابحث في ذاكرة Kiven الدائمة عن معلومة. سيب query فاضي عشان تجيب كل الذاكرة. "
            "استخدمها لما المستخدم يسأل 'إنت فاكر...؟' أو 'إيه اللي تعرفه عني؟'"
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "query": {"type": "string", "description": "نص للبحث (فارغ = كل الذاكرة)."},
                "limit": {"type": "integer", "minimum": 1, "maximum": 50, "description": "عدد النتائج. افتراضي 10."},
            },
        },
        "handler": recall,
    },
    "forget": {
        "name": "forget",
        "description": "امسح معلومة من الذاكرة بالـ ID بتاعها. استخدم recall الأول عشان تعرف الـ IDs.",
        "input_schema": {
            "type": "object",
            "properties": {
                "fact_id": {"type": "integer", "description": "رقم المعلومة (يظهر في recall)."},
            },
            "required": ["fact_id"],
        },
        "handler": forget,
    },
    "get_current_time": {
        "name": "get_current_time",
        "description": "اجلب الوقت والتاريخ الحاليين. استخدمها لما المستخدم يسأل 'إيه الساعة' أو 'النهارده إيه'.",
        "input_schema": {
            "type": "object",
            "properties": {
                "timezone": {"type": "string", "description": "Timezone IANA name. Default: Africa/Cairo."},
            },
        },
        "handler": get_current_time,
    },
    "get_weather": {
        "name": "get_weather",
        "description": (
            "اجلب حالة الطقس الحالية لمدينة. الـ city يفضّل يتكتب بالإنجليزية للدقة "
            "(مثلاً 'Cairo' بدل 'القاهرة'). استخدمها لما المستخدم يسأل عن الجو، الحرارة، الطقس."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "city": {"type": "string", "description": "اسم المدينة (English preferred). مثل: Cairo, Alexandria, Riyadh."},
            },
            "required": ["city"],
        },
        "handler": get_weather,
    },
    "calculate": {
        "name": "calculate",
        "description": (
            "احسب معادلة رياضية بدقة. يدعم +, -, *, /, %, ** (أس). "
            "استخدمها لما المستخدم يطلب حساب أرقام كبيرة، نسب، فواتير، إلخ — "
            "بدل ما تحسب في عقلك (احتمال خطأ)."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "expression": {"type": "string", "description": "Math expression, e.g. '15% * 2500' or '(45+67)*3.14'."},
            },
            "required": ["expression"],
        },
        "handler": calculate,
    },
}

def list_claude_tools():
    return [
        {"name": t["name"], "description": t["description"], "input_schema": t["input_schema"]}
        for t in TOOLS_REGISTRY.values()
    ]

NEWS_KEYWORDS_AR = ["أخبار", "اخبار", "خبر", "إخبار", "إيه اللي بيحصل", "النهارده", "اليوم", "آخر", "احدث", "أحدث", "حاليا", "حالياً"]
NEWS_KEYWORDS_EN = ["news", "today", "latest", "current", "recent", "headlines", "happening"]

IMAGE_KEYWORDS_AR = ["صورة", "صوره", "ارسم", "ارسمي", "اعمل صورة", "اعملي صورة", "كرت", "تصميم", "رسمة", "لوحة", "خلفية", "بوستر", "اعمل لي صورة", "اعمل كرت", "صمم", "صممي"]
IMAGE_KEYWORDS_EN = ["image", "picture", "draw", "design", "card", "poster", "wallpaper", "illustration", "render"]

TIME_KEYWORDS_AR = ["الساعة كام", "ساعة كام", "إيه الوقت", "ايه الوقت", "النهارده إيه", "النهارده ايه", "إيه التاريخ", "التاريخ", "اليوم إيه", "اليوم ايه"]
TIME_KEYWORDS_EN = ["what time", "what's the time", "current time", "today's date", "what day", "what date"]

WEATHER_KEYWORDS_AR = ["الجو", "الطقس", "الحرارة", "حرارة", "بتمطر", "حر", "برد", "رطوبة"]
WEATHER_KEYWORDS_EN = ["weather", "temperature", "raining", "humidity", "forecast"]

CALC_KEYWORDS_AR = ["احسب", "احسبلي", "حاسبلي", "كام يساوي", "كام بيطلع", "اضرب", "اقسم", "نسبة"]
CALC_KEYWORDS_EN = ["calculate", "compute", "what's", "how much is"]

def detect_time_intent(message: str) -> bool:
    low = message.lower()
    return any(k in message for k in TIME_KEYWORDS_AR) or any(k in low for k in TIME_KEYWORDS_EN)

def detect_weather_intent(message: str) -> bool:
    low = message.lower()
    return any(k in message for k in WEATHER_KEYWORDS_AR) or any(k in low for k in WEATHER_KEYWORDS_EN)

def detect_calc_intent(message: str) -> bool:
    low = message.lower()
    has_kw = any(k in message for k in CALC_KEYWORDS_AR) or any(k in low for k in CALC_KEYWORDS_EN)
    # Or message looks like a math expression
    import re as _re
    looks_mathy = bool(_re.search(r"\d+\s*[\+\-\*\/×÷%]\s*\d+", message))
    return has_kw or looks_mathy

def detect_search_intent(message: str) -> tuple[bool, str]:
    """Heuristic: returns (should_search, kind) for use with non-tool-calling providers (Gemini)."""
    low = message.lower()
    has_ar = any(k in message for k in NEWS_KEYWORDS_AR)
    has_en = any(k in low for k in NEWS_KEYWORDS_EN)
    if has_ar or has_en:
        return True, "news"
    return False, "web"

def detect_image_intent(message: str) -> bool:
    low = message.lower()
    return any(k in message for k in IMAGE_KEYWORDS_AR) or any(k in low for k in IMAGE_KEYWORDS_EN)

SYSTEM_PROMPT = f"""أنت {ASSISTANT_NAME_AR} ({ASSISTANT_NAME})، مساعد ذكي شخصي شامل لعمرو.

عن عمرو:
- يعمل في مجال IT في شركة Egyptian Cement
- يستخدم SAP، Microsoft 365، SharePoint
- يتكلم العربية المصرية والإنجليزية

🎯 إنت مساعد شخصي عام — مش محدود بأي تخصص:
- بتساعد في الكتابة الإبداعية: كروت معايدة، رسائل، خطابات، أدعية، أشعار، نصوص رسمية، إيميلات
- بتساعد في المعلومات والتقنيات: Claude Skills، Plugins، APIs، تطوير، IT
- بتساعد في الحياة اليومية: تنظيم، تخطيط، نصائح، أفكار
- بتساعد في الأخبار والمعلومات الجارية (عبر web_search)
- بتساعد في أي حاجة تانية عمرو يطلبها — متترددش

❌ ممنوع تقول "أنا مش قادر" أو "إمكانياتي محدودة" أو "أنا متخصص فقط في..." لأي طلب نصي عادي.
- لو طلب كرت معايدة → اكتب نص الكرت بشكل جميل ومنسق بالـ Emojis والتنسيق المناسب
- لو طلب دعاء → اكتب الدعاء كامل ومنسق
- لو طلب يوقع باسم معين → اكتب التوقيع كما طلب
- إنت مش بترسم صور (مش متاح حالياً)، لكن النصوص الإبداعية والمنسقة بالـ Emojis والـ Unicode متاحة لك
- استخدم **Markdown** فقط للتنسيق (## ، **bold** ، --- ، •) و **Emojis** بدل HTML tags. الـ UI نصي بحت ومش هيـ render أي HTML

📝 أسلوب الرد:
- ترد دائماً بالعربية المصرية الواضحة (مع الفصحى عند الحاجة زي الأدعية والآيات)
- للأسئلة العادية: 2-3 أسطر مختصرة
- للطلبات الإبداعية (كروت، أدعية، خطابات): رد كامل وجميل ومنسق — مفيش حد أقصى
- لو السؤال إنجليزي، رد بالعربية ما لم يُطلب الإنجليزية صراحة
- اسمك {ASSISTANT_NAME_AR} مش Jarvis

🔧 الأدوات المتاحة:
- **web_search**: استخدمها فوراً وبدون تردد لأي معلومة حديثة من النت (أخبار، أسعار، نتائج، أحداث). متقولش "مش قادر" — استخدم الأداة. اذكر المصادر باختصار.
- **generate_image**: للصور والكروت والرسومات والتصاميم. الـ prompt إنجليزي مفصل. اعرض الـ Markdown image في ردك + النص العربي.
- **get_current_time**: للوقت والتاريخ الحالي (Cairo افتراضياً). استخدمها بدل ما تخمن.
- **get_weather**: لحالة الطقس. اسم المدينة بالإنجليزية (Cairo, Alexandria, ...).
- **calculate**: لأي حسابات رياضية. متحسبش في عقلك — استخدم الأداة عشان الدقة.
- **remember(fact)**: احفظ معلومات شخصية مهمة لما المستخدم يقول "افتكر إن..." أو يدّيك معلومة عن نفسه. الذاكرة بتفضل بين الجلسات.
- **recall(query)**: ابحث في ذاكرتك القديمة. استخدمها لما المستخدم يسأل "إنت فاكر...؟" أو "إيه اللي تعرفه عني؟".
- **forget(fact_id)**: امسح معلومة بعد ما المستخدم يطلب.

{knowledge_base.get('summary', '')}

استخدم قاعدة المعرفة دي للأسئلة المتعلقة بـ Claude Skills والـ Plugins، واستخدم web_search للأخبار والمعلومات الحديثة، وردّ بثقة وإبداع على أي طلب تاني."""

def build_system_prompt() -> str:
    """Return SYSTEM_PROMPT augmented with current persistent memory facts."""
    base = SYSTEM_PROMPT
    mem_block = memory_summary_for_prompt()
    if mem_block:
        return f"{base}\n\n{mem_block}\n\n(لو المستخدم سأل عن نفسه أو معلومات سابقة، استخدم الذاكرة دي مباشرة بدون recall.)"
    return base

# ============================================
# EMBEDDED HTML UI
# ============================================
HTML_UI = f"""<!DOCTYPE html>
<html lang="ar" dir="rtl">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0, user-scalable=no">
    <title>{ASSISTANT_NAME_AR} | المساعد الذكي</title>
    <link href="https://fonts.googleapis.com/css2?family=Cairo:wght@400;700;900&display=swap" rel="stylesheet">
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; -webkit-tap-highlight-color: transparent; }}
        :root {{
            --bg-deep: #050d1a; --bg-card: #0e1a2e; --accent: #00d4ff;
            --accent-glow: rgba(0, 212, 255, 0.4); --accent-warm: #ff6b35;
            --text: #e8f4ff; --text-dim: #7a9cc6; --success: #00ff88;
        }}
        body {{ font-family: 'Cairo', sans-serif; background: var(--bg-deep); color: var(--text); min-height: 100vh; overflow: hidden; }}
        body::before {{ content: ''; position: fixed; inset: 0; background: radial-gradient(circle at 20% 30%, rgba(0,212,255,0.08) 0%, transparent 50%), radial-gradient(circle at 80% 70%, rgba(255,107,53,0.05) 0%, transparent 50%); pointer-events: none; z-index: 0; }}
        .container {{ position: relative; z-index: 1; height: 100vh; display: flex; flex-direction: column; max-width: 600px; margin: 0 auto; }}
        .header {{ padding: 16px 20px; background: rgba(14,26,46,0.8); backdrop-filter: blur(20px); border-bottom: 1px solid rgba(0,212,255,0.2); display: flex; align-items: center; justify-content: space-between; gap: 12px; flex-wrap: wrap; }}
        .logo-section {{ display: flex; align-items: center; gap: 12px; flex: 1; min-width: 0; }}
        .arc-reactor {{ width: 40px; height: 40px; border-radius: 50%; background: radial-gradient(circle, var(--accent) 0%, rgba(0,212,255,0.2) 70%); box-shadow: 0 0 20px var(--accent-glow), inset 0 0 10px rgba(255,255,255,0.3); animation: pulse 3s ease-in-out infinite; position: relative; flex-shrink: 0; }}
        .arc-reactor::after {{ content: ''; position: absolute; inset: 8px; border-radius: 50%; background: var(--accent); box-shadow: 0 0 8px white; }}
        @keyframes pulse {{ 0%,100% {{ box-shadow: 0 0 20px var(--accent-glow), inset 0 0 10px rgba(255,255,255,0.3); }} 50% {{ box-shadow: 0 0 30px var(--accent-glow), 0 0 50px rgba(0,212,255,0.2); }} }}
        .title h1 {{ font-size: 18px; font-weight: 700; }}
        .title .subtitle {{ font-size: 11px; color: var(--text-dim); display: flex; align-items: center; gap: 6px; }}
        .status-dot {{ width: 8px; height: 8px; border-radius: 50%; background: var(--success); box-shadow: 0 0 8px var(--success); animation: blink 2s infinite; flex-shrink: 0; }}
        @keyframes blink {{ 0%,100% {{ opacity: 1; }} 50% {{ opacity: 0.4; }} }}

        /* Controls bar - moved to TOP, replaces old provider-bar and header-actions */
        .controls-bar {{ display: flex; gap: 6px; align-items: center; flex-shrink: 0; }}
        .provider-btn {{ padding: 8px 12px; border-radius: 10px; background: rgba(0,212,255,0.05); border: 1px solid rgba(0,212,255,0.15); color: var(--text-dim); font-family: inherit; font-size: 12px; font-weight: 600; cursor: pointer; }}
        .provider-btn.active {{ background: rgba(0,212,255,0.15); border-color: var(--accent); color: var(--accent); box-shadow: 0 0 12px rgba(0,212,255,0.2); }}
        .provider-btn:disabled {{ opacity: 0.4; cursor: not-allowed; }}
        .icon-btn {{ width: 36px; height: 36px; border-radius: 10px; background: rgba(0,212,255,0.1); border: 1px solid rgba(0,212,255,0.2); color: var(--accent); cursor: pointer; display: flex; align-items: center; justify-content: center; font-size: 16px; }}

        .chat-area {{ flex: 1; overflow-y: auto; padding: 20px; display: flex; flex-direction: column; gap: 14px; }}
        .message {{ max-width: 85%; padding: 12px 16px; border-radius: 18px; font-size: 15px; line-height: 1.6; word-wrap: break-word; }}
        .message.user {{ align-self: flex-start; background: linear-gradient(135deg, #00d4ff 0%, #0088cc 100%); color: white; border-bottom-right-radius: 4px; box-shadow: 0 4px 12px rgba(0,212,255,0.2); }}
        .message.bot {{ align-self: flex-end; background: linear-gradient(135deg, #1a2e4e 0%, #0e1a2e 100%); color: var(--text); border: 1px solid rgba(0,212,255,0.15); border-bottom-left-radius: 4px; white-space: pre-wrap; }}
        .welcome {{ text-align: center; padding: 32px 20px; color: var(--text-dim); }}
        .welcome h2 {{ color: var(--accent); font-size: 24px; margin-bottom: 8px; font-weight: 900; letter-spacing: 0.5px; }}
        .welcome p {{ font-size: 14px; line-height: 1.8; margin-top: 8px; }}
        .welcome p strong {{ color: var(--text); }}
        .welcome .hint {{ font-size: 12px; color: var(--text-dim); margin-top: 16px; opacity: 0.8; }}
        .caps {{ display: flex; flex-wrap: wrap; gap: 6px; justify-content: center; margin: 16px 0; }}
        .cap {{ font-size: 11px; padding: 5px 10px; border-radius: 999px; background: rgba(0,212,255,0.08); border: 1px solid rgba(0,212,255,0.25); color: var(--accent); white-space: nowrap; }}
        .typing {{ align-self: flex-end; display: flex; gap: 5px; padding: 14px 18px; background: linear-gradient(135deg, #1a2e4e 0%, #0e1a2e 100%); border-radius: 18px; border: 1px solid rgba(0,212,255,0.15); }}
        .typing-dot {{ width: 8px; height: 8px; border-radius: 50%; background: var(--accent); animation: typing 1.4s infinite; }}
        .typing-dot:nth-child(2) {{ animation-delay: 0.2s; }}
        .typing-dot:nth-child(3) {{ animation-delay: 0.4s; }}
        @keyframes typing {{ 0%,60%,100% {{ transform: translateY(0); opacity: 0.5; }} 30% {{ transform: translateY(-8px); opacity: 1; }} }}
        .input-area {{ padding: 14px 16px 20px; background: rgba(14,26,46,0.95); border-top: 1px solid rgba(0,212,255,0.2); }}
        .input-row {{ display: flex; gap: 8px; align-items: flex-end; }}
        .text-input {{ flex: 1; padding: 12px 16px; background: rgba(0,212,255,0.05); border: 1px solid rgba(0,212,255,0.2); border-radius: 22px; color: var(--text); font-family: inherit; font-size: 15px; outline: none; resize: none; max-height: 100px; min-height: 44px; }}
        .text-input:focus {{ border-color: var(--accent); }}
        .send-btn, .mic-btn {{ width: 44px; height: 44px; border-radius: 50%; border: none; cursor: pointer; display: flex; align-items: center; justify-content: center; font-size: 18px; flex-shrink: 0; }}
        .send-btn {{ background: linear-gradient(135deg, #00d4ff 0%, #0088cc 100%); color: white; box-shadow: 0 4px 12px rgba(0,212,255,0.3); }}
        .mic-btn {{ background: rgba(255,107,53,0.15); border: 2px solid var(--accent-warm); color: var(--accent-warm); }}
        .mic-btn.recording {{ background: var(--accent-warm); color: white; box-shadow: 0 0 20px rgba(255,107,53,0.6); animation: recordPulse 1s infinite; }}
        @keyframes recordPulse {{ 0%,100% {{ box-shadow: 0 0 20px rgba(255,107,53,0.6); }} 50% {{ box-shadow: 0 0 30px rgba(255,107,53,0.9); }} }}

        /* TOAST - top-left corner, doesn't cover anything */
        .toast {{ position: fixed; top: 12px; left: 12px; max-width: 280px; padding: 10px 14px; background: var(--bg-card); border: 1px solid var(--accent); border-radius: 10px; color: var(--text); font-size: 13px; transform: translateX(calc(-100% - 20px)); transition: transform 0.3s; z-index: 9999; box-shadow: 0 8px 24px rgba(0,0,0,0.5); }}
        .toast.show {{ transform: translateX(0); }}
        .toast.error {{ border-color: var(--accent-warm); }}

        .kb-badge {{ font-size: 10px; padding: 2px 6px; border-radius: 6px; background: rgba(0,255,136,0.15); color: var(--success); border: 1px solid rgba(0,255,136,0.3); }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <div class="logo-section">
                <div class="arc-reactor"></div>
                <div class="title">
                    <h1>{ASSISTANT_NAME_AR}</h1>
                    <div class="subtitle">
                        <span class="status-dot"></span>
                        <span id="status-text">جاهز للخدمة</span>
                        <span class="kb-badge" id="kb-badge" style="display:none">📚 KB</span>
                    </div>
                </div>
            </div>
            <div class="controls-bar">
                <button class="provider-btn active" id="btn-claude" onclick="switchProvider('claude')">🧠 Claude</button>
                <button class="provider-btn" id="btn-gemini" onclick="switchProvider('gemini')">✨ Gemini</button>
                <button class="icon-btn" onclick="clearChat()" title="مسح">🗑️</button>
                <button class="icon-btn" onclick="doLogout()" title="خروج">🚪</button>
            </div>
        </div>
        <div class="chat-area" id="chat-area">
            <div class="welcome">
                <h2>أهلاً يا عمرو 👋</h2>
                <p>أنا <strong>{ASSISTANT_NAME_AR}</strong> ({ASSISTANT_NAME} v{ASSISTANT_VERSION})، مساعدك الشخصي الذكي.</p>
                <div class="caps">
                    <span class="cap">🌐 بحث</span>
                    <span class="cap">🎨 صور</span>
                    <span class="cap">🧠 ذاكرة</span>
                    <span class="cap">🕒 وقت</span>
                    <span class="cap">🌤️ طقس</span>
                    <span class="cap">🧮 حساب</span>
                </div>
                <p class="hint">اكتب أو اضغط 🎤 وكلّمني بالعربي. جرّب: "اعمل لي كرت معايدة" أو "إيه أخبار النهارده"</p>
            </div>
        </div>
        <div class="input-area">
            <div class="input-row">
                <button class="mic-btn" id="mic-btn" onclick="toggleRecording()">🎤</button>
                <textarea class="text-input" id="text-input" placeholder="اكتب رسالتك..." rows="1" onkeydown="handleKey(event)"></textarea>
                <button class="send-btn" onclick="sendMessage()">➤</button>
            </div>
        </div>
    </div>
    <div class="toast" id="toast"></div>
    <script>
        const API_BASE = window.location.origin;
        let currentProvider = 'gemini';
        let mediaRecorder = null;
        let audioChunks = [];
        let isRecording = false;
        let conversationStarted = false;
        let recordingStartTime = 0;
        let recordingTimer = null;

        window.addEventListener('load', async () => {{
            await checkStatus();
            const input = document.getElementById('text-input');
            input.addEventListener('input', () => {{ input.style.height = 'auto'; input.style.height = Math.min(input.scrollHeight, 100) + 'px'; }});
        }});

        async function checkStatus() {{
            try {{
                const res = await fetch(`${{API_BASE}}/api/status`);
                if (res.status === 401) {{ window.location.href = '/login'; return; }}
                const data = await res.json();
                if (!data.claude_available) document.getElementById('btn-claude').disabled = true;
                if (!data.gemini_available) document.getElementById('btn-gemini').disabled = true;
                if (!data.claude_available && data.gemini_available) switchProvider('gemini');
                else if (data.claude_available && !data.gemini_available) switchProvider('claude');
                if (data.knowledge_base_loaded > 0) {{
                    const badge = document.getElementById('kb-badge');
                    badge.style.display = 'inline-block';
                    badge.textContent = `📚 ${{data.knowledge_base_loaded}} ملف`;
                }}
            }} catch (err) {{
                showToast('❌ خطأ في الاتصال', true);
            }}
        }}

        function switchProvider(provider) {{
            if (document.getElementById(`btn-${{provider}}`).disabled) {{
                showToast(`${{provider}} غير متاح`, true);
                return;
            }}
            currentProvider = provider;
            document.querySelectorAll('.provider-btn').forEach(b => b.classList.remove('active'));
            document.getElementById(`btn-${{provider}}`).classList.add('active');
            showToast(`✅ ${{provider}}`);
        }}

        function handleKey(event) {{
            if (event.key === 'Enter' && !event.shiftKey) {{ event.preventDefault(); sendMessage(); }}
        }}

        async function sendMessage() {{
            const input = document.getElementById('text-input');
            const message = input.value.trim();
            if (!message) return;
            input.value = '';
            input.style.height = 'auto';
            if (!conversationStarted) {{ document.querySelector('.welcome')?.remove(); conversationStarted = true; }}
            addMessage('user', message);
            const typing = showTyping();
            try {{
                const res = await fetch(`${{API_BASE}}/api/chat`, {{
                    method: 'POST',
                    headers: {{'Content-Type': 'application/json'}},
                    body: JSON.stringify({{message, provider: currentProvider}})
                }});
                typing.remove();
                if (res.status === 401) {{ window.location.href = '/login'; return; }}
                if (!res.ok) {{ const err = await res.json(); throw new Error(err.detail || 'خطأ'); }}
                const data = await res.json();
                addMessage('bot', data.response);
                speakText(data.response);
            }} catch (err) {{
                typing.remove();
                addMessage('bot', `❌ خطأ: ${{err.message}}`);
            }}
        }}

        function safeUrl(url) {{
            return /^https?:\\/\\//i.test(url) ? url : '#';
        }}
        function escapeHtml(s) {{
            return s.replace(/&/g, '&amp;').replace(/</g, '&lt;').replace(/>/g, '&gt;');
        }}
        function renderMarkdown(text) {{
            let html = escapeHtml(text);
            // Code fences
            html = html.replace(/```([\\s\\S]*?)```/g, (_, code) =>
                `<pre style="background:rgba(0,0,0,0.3);padding:10px;border-radius:8px;overflow-x:auto;direction:ltr;text-align:left;font-size:13px"><code>${{code}}</code></pre>`);
            // Images: ![alt](url) - MUST come before links
            html = html.replace(/!\\[([^\\]]*)\\]\\((https?:\\/\\/[^\\s)]+)\\)/g, (_, alt, url) =>
                `<img src="${{safeUrl(url)}}" alt="${{escapeHtml(alt)}}" loading="lazy" style="max-width:100%;border-radius:14px;margin:10px 0;display:block;box-shadow:0 6px 20px rgba(0,0,0,0.35)">`);
            // Links: [text](url)
            html = html.replace(/\\[([^\\]]+)\\]\\((https?:\\/\\/[^\\s)]+)\\)/g, (_, t, url) =>
                `<a href="${{safeUrl(url)}}" target="_blank" rel="noopener noreferrer" style="color:var(--accent);text-decoration:underline">${{t}}</a>`);
            // Bold
            html = html.replace(/\\*\\*([^*\\n]+)\\*\\*/g, '<strong>$1</strong>');
            // Italic (single * not part of **)
            html = html.replace(/(^|[^*])\\*([^*\\n]+)\\*(?!\\*)/g, '$1<em>$2</em>');
            // Headers
            html = html.replace(/^### (.+)$/gm, '<h3 style="margin:8px 0;font-size:1.05em;color:var(--accent)">$1</h3>');
            html = html.replace(/^## (.+)$/gm, '<h2 style="margin:10px 0;font-size:1.15em;color:var(--accent)">$1</h2>');
            html = html.replace(/^# (.+)$/gm, '<h1 style="margin:12px 0;font-size:1.25em;color:var(--accent)">$1</h1>');
            // Horizontal rule
            html = html.replace(/^---+$/gm, '<hr style="border:none;border-top:1px solid rgba(0,212,255,0.25);margin:12px 0">');
            // Line breaks
            html = html.replace(/\\n/g, '<br>');
            return html;
        }}

        function addMessage(type, text) {{
            const chat = document.getElementById('chat-area');
            const msg = document.createElement('div');
            msg.className = `message ${{type}}`;
            if (type === 'bot') {{
                msg.innerHTML = renderMarkdown(text);
            }} else {{
                msg.textContent = text;
            }}
            chat.appendChild(msg);
            chat.scrollTop = chat.scrollHeight;
        }}

        function showTyping() {{
            const chat = document.getElementById('chat-area');
            const typing = document.createElement('div');
            typing.className = 'typing';
            typing.innerHTML = '<div class="typing-dot"></div><div class="typing-dot"></div><div class="typing-dot"></div>';
            chat.appendChild(typing);
            chat.scrollTop = chat.scrollHeight;
            return typing;
        }}

        async function speakText(text) {{
            try {{
                const cleanText = text
                    .replace(/!\\[[^\\]]*\\]\\([^)]+\\)/g, ' ')        // strip image markdown
                    .replace(/\\[([^\\]]+)\\]\\([^)]+\\)/g, '$1')      // strip link URLs, keep text
                    .replace(/```[\\s\\S]*?```/g, ' ')                  // strip code blocks
                    .replace(/^#+\\s+/gm, '')                           // strip header markers
                    .replace(/^---+$/gm, ' ')                            // strip rules
                    .replace(/[*_`#\\[\\]]/g, '');                       // strip remaining markdown chars
                const res = await fetch(`${{API_BASE}}/api/speak`, {{
                    method: 'POST',
                    headers: {{'Content-Type': 'application/json'}},
                    body: JSON.stringify({{text: cleanText.substring(0, 500), voice: 'shakir_egypt'}})
                }});
                if (!res.ok) return;
                const blob = await res.blob();
                const audio = new Audio(URL.createObjectURL(blob));
                audio.play();
            }} catch (err) {{ console.error('TTS error:', err); }}
        }}

        async function toggleRecording() {{
            if (isRecording) stopRecording();
            else await startRecording();
        }}

        async function startRecording() {{
            try {{
                const stream = await navigator.mediaDevices.getUserMedia({{
                    audio: {{
                        echoCancellation: true,
                        noiseSuppression: true,
                        autoGainControl: true,
                        sampleRate: 44100
                    }}
                }});

                // Pick best supported MIME type
                let mimeType = 'audio/webm;codecs=opus';
                if (!MediaRecorder.isTypeSupported(mimeType)) {{
                    mimeType = 'audio/webm';
                    if (!MediaRecorder.isTypeSupported(mimeType)) {{
                        mimeType = 'audio/mp4';
                        if (!MediaRecorder.isTypeSupported(mimeType)) {{
                            mimeType = '';  // Browser default
                        }}
                    }}
                }}

                mediaRecorder = mimeType ? new MediaRecorder(stream, {{mimeType}}) : new MediaRecorder(stream);
                audioChunks = [];
                recordingStartTime = Date.now();

                mediaRecorder.ondataavailable = (e) => {{
                    if (e.data && e.data.size > 0) audioChunks.push(e.data);
                }};

                mediaRecorder.onstop = async () => {{
                    const duration = (Date.now() - recordingStartTime) / 1000;
                    const blob = new Blob(audioChunks, {{type: mediaRecorder.mimeType || 'audio/webm'}});

                    stream.getTracks().forEach(t => t.stop());

                    if (duration < 1.5) {{
                        showToast(`⚠️ تسجيل قصير (${{duration.toFixed(1)}}s)`, true);
                        document.getElementById('status-text').textContent = 'جاهز للخدمة';
                        return;
                    }}

                    if (blob.size < 2000) {{
                        showToast(`⚠️ تسجيل فارغ (${{blob.size}}b)`, true);
                        document.getElementById('status-text').textContent = 'جاهز للخدمة';
                        return;
                    }}

                    await transcribeAndSend(blob);
                }};

                mediaRecorder.start(100);  // Collect data every 100ms
                isRecording = true;
                document.getElementById('mic-btn').classList.add('recording');
                startRecordingTimer();
            }} catch (err) {{
                showToast('❌ مفيش صلاحية للميكروفون', true);
                console.error('Mic error:', err);
            }}
        }}

        function startRecordingTimer() {{
            const statusEl = document.getElementById('status-text');
            recordingTimer = setInterval(() => {{
                if (!isRecording) {{ clearInterval(recordingTimer); return; }}
                const elapsed = ((Date.now() - recordingStartTime) / 1000).toFixed(1);
                statusEl.textContent = `🔴 ${{elapsed}}s`;
            }}, 100);
        }}

        function stopRecording() {{
            if (mediaRecorder && isRecording) {{
                mediaRecorder.stop();
                isRecording = false;
                if (recordingTimer) clearInterval(recordingTimer);
                document.getElementById('mic-btn').classList.remove('recording');
                document.getElementById('status-text').textContent = 'جاري المعالجة...';
            }}
        }}

        async function transcribeAndSend(blob) {{
            const formData = new FormData();
            const ext = blob.type.includes('mp4') ? '.mp4' : '.webm';
            formData.append('audio', blob, `recording${{ext}}`);

            try {{
                const res = await fetch(`${{API_BASE}}/api/transcribe`, {{ method: 'POST', body: formData }});
                document.getElementById('status-text').textContent = 'جاهز للخدمة';
                if (!res.ok) {{
                    const err = await res.json();
                    throw new Error(err.detail || 'فشل التحويل');
                }}
                const data = await res.json();
                if (data.text && data.text.length > 1) {{
                    document.getElementById('text-input').value = data.text;
                    await sendMessage();
                }} else {{
                    showToast('⚠️ مسمعتش حاجة', true);
                }}
            }} catch (err) {{
                document.getElementById('status-text').textContent = 'جاهز للخدمة';
                showToast(`❌ ${{err.message}}`, true);
            }}
        }}

        async function doLogout() {{
            if (!confirm('تسجيل خروج؟')) return;
            try {{
                await fetch(`${{API_BASE}}/api/logout`, {{method: 'POST'}});
            }} catch (err) {{}}
            window.location.href = '/login';
        }}

        async function clearChat() {{
            if (!confirm('متأكد؟')) return;
            try {{
                await fetch(`${{API_BASE}}/api/clear_history`, {{method: 'POST'}});
                document.getElementById('chat-area').innerHTML = '<div class="welcome"><h2>اتمسحت ✨</h2><p>ابدأ من جديد</p></div>';
                conversationStarted = false;
                showToast('✅ تم المسح');
            }} catch (err) {{ showToast('❌ خطأ', true); }}
        }}

        function showToast(msg, isError = false) {{
            const toast = document.getElementById('toast');
            toast.textContent = msg;
            toast.classList.toggle('error', isError);
            toast.classList.add('show');
            setTimeout(() => toast.classList.remove('show'), 2200);
        }}
    </script>
</body>
</html>"""

# ============================================
# LOGIN PAGE HTML
# ============================================
LOGIN_HTML = f"""<!DOCTYPE html>
<html lang="ar" dir="rtl">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>🔒 {ASSISTANT_NAME_AR} | تسجيل الدخول</title>
    <link href="https://fonts.googleapis.com/css2?family=Cairo:wght@400;700;900&display=swap" rel="stylesheet">
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        :root {{
            --bg-deep: #050d1a; --bg-card: #0e1a2e; --accent: #00d4ff;
            --accent-glow: rgba(0, 212, 255, 0.4); --accent-warm: #ff6b35;
            --text: #e8f4ff; --text-dim: #7a9cc6; --danger: #ff3860;
        }}
        body {{
            font-family: 'Cairo', sans-serif; background: var(--bg-deep);
            color: var(--text); min-height: 100vh; display: flex;
            align-items: center; justify-content: center; padding: 20px;
        }}
        body::before {{
            content: ''; position: fixed; inset: 0;
            background: radial-gradient(circle at 20% 30%, rgba(0,212,255,0.08) 0%, transparent 50%),
                        radial-gradient(circle at 80% 70%, rgba(255,107,53,0.05) 0%, transparent 50%);
            pointer-events: none; z-index: 0;
        }}
        .login-card {{
            position: relative; z-index: 1; max-width: 420px; width: 100%;
            background: rgba(14,26,46,0.95); backdrop-filter: blur(20px);
            border: 1px solid rgba(0,212,255,0.3); border-radius: 24px;
            padding: 40px 32px; box-shadow: 0 20px 60px rgba(0,0,0,0.5), 0 0 60px rgba(0,212,255,0.1);
        }}
        .arc-reactor {{
            width: 80px; height: 80px; margin: 0 auto 20px; border-radius: 50%;
            background: radial-gradient(circle, var(--accent) 0%, rgba(0,212,255,0.2) 70%);
            box-shadow: 0 0 40px var(--accent-glow), inset 0 0 20px rgba(255,255,255,0.3);
            animation: pulse 3s ease-in-out infinite; position: relative;
        }}
        .arc-reactor::after {{
            content: ''; position: absolute; inset: 18px; border-radius: 50%;
            background: var(--accent); box-shadow: 0 0 12px white;
        }}
        @keyframes pulse {{
            0%,100% {{ box-shadow: 0 0 40px var(--accent-glow), inset 0 0 20px rgba(255,255,255,0.3); }}
            50% {{ box-shadow: 0 0 60px var(--accent-glow), 0 0 100px rgba(0,212,255,0.2); }}
        }}
        h1 {{ text-align: center; font-size: 26px; margin-bottom: 4px; }}
        .subtitle {{ text-align: center; color: var(--text-dim); font-size: 14px; margin-bottom: 32px; }}
        .form-group {{ margin-bottom: 20px; }}
        label {{ display: block; font-size: 13px; color: var(--text-dim); margin-bottom: 8px; }}
        input[type="password"] {{
            width: 100%; padding: 14px 16px; background: rgba(0,212,255,0.05);
            border: 1px solid rgba(0,212,255,0.2); border-radius: 12px;
            color: var(--text); font-family: inherit; font-size: 16px;
            outline: none; transition: all 0.2s; direction: ltr;
        }}
        input[type="password"]:focus {{
            border-color: var(--accent); box-shadow: 0 0 0 3px rgba(0,212,255,0.15);
        }}
        button {{
            width: 100%; padding: 14px; background: linear-gradient(135deg, #00d4ff 0%, #0088cc 100%);
            color: white; border: none; border-radius: 12px; font-family: inherit;
            font-size: 16px; font-weight: 700; cursor: pointer;
            box-shadow: 0 4px 16px rgba(0,212,255,0.3); transition: transform 0.1s;
        }}
        button:hover {{ transform: translateY(-1px); box-shadow: 0 6px 20px rgba(0,212,255,0.4); }}
        button:active {{ transform: translateY(0); }}
        button:disabled {{ opacity: 0.6; cursor: not-allowed; }}
        .error-msg {{
            display: none; margin-top: 16px; padding: 12px;
            background: rgba(255,56,96,0.1); border: 1px solid var(--danger);
            border-radius: 10px; color: var(--danger); font-size: 14px; text-align: center;
        }}
        .error-msg.show {{ display: block; }}
        .hint {{ margin-top: 24px; text-align: center; font-size: 12px; color: var(--text-dim); }}
    </style>
</head>
<body>
    <div class="login-card">
        <div class="arc-reactor"></div>
        <h1>{ASSISTANT_NAME_AR}</h1>
        <div class="subtitle">المساعد الشخصي - مطلوب كلمة المرور</div>
        <form id="login-form" onsubmit="return doLogin(event)">
            <div class="form-group">
                <label for="password">كلمة المرور</label>
                <input type="password" id="password" name="password" required autofocus
                       placeholder="••••••••••••" autocomplete="current-password">
            </div>
            <button type="submit" id="submit-btn">دخول 🔓</button>
            <div class="error-msg" id="error-msg"></div>
        </form>
        <div class="hint">🔒 الوصول مقصور على المالك</div>
    </div>
    <script>
        async function doLogin(e) {{
            e.preventDefault();
            const password = document.getElementById('password').value;
            const errorEl = document.getElementById('error-msg');
            const btn = document.getElementById('submit-btn');
            errorEl.classList.remove('show');
            btn.disabled = true;
            btn.textContent = 'جاري التحقق...';
            try {{
                const res = await fetch('/api/login', {{
                    method: 'POST',
                    headers: {{'Content-Type': 'application/json'}},
                    body: JSON.stringify({{password}})
                }});
                if (res.ok) {{
                    window.location.href = '/';
                }} else {{
                    const data = await res.json().catch(() => ({{detail: 'كلمة مرور خاطئة'}}));
                    errorEl.textContent = '❌ ' + (data.detail || 'كلمة مرور خاطئة');
                    errorEl.classList.add('show');
                    document.getElementById('password').value = '';
                    document.getElementById('password').focus();
                }}
            }} catch (err) {{
                errorEl.textContent = '❌ خطأ في الاتصال';
                errorEl.classList.add('show');
            }} finally {{
                btn.disabled = false;
                btn.textContent = 'دخول 🔓';
            }}
            return false;
        }}
    </script>
</body>
</html>"""

# ============================================
# SESSION / AUTH HELPERS
# ============================================
def _sign(message: str) -> str:
    mac = hmac.new(KEVIN_SESSION_SECRET.encode("utf-8"),
                   message.encode("utf-8"), hashlib.sha256).digest()
    return base64.urlsafe_b64encode(mac).decode("ascii").rstrip("=")

def make_session_token() -> str:
    """Create signed session token: <timestamp>.<signature>"""
    issued = str(int(datetime.now().timestamp()))
    sig = _sign(issued)
    return f"{issued}.{sig}"

def verify_session_token(token: str) -> bool:
    if not token or "." not in token:
        return False
    try:
        issued, sig = token.rsplit(".", 1)
        expected = _sign(issued)
        if not hmac.compare_digest(sig, expected):
            return False
        issued_ts = int(issued)
        age = int(datetime.now().timestamp()) - issued_ts
        return 0 <= age <= SESSION_TTL_SECONDS
    except Exception:
        return False

def require_auth(kevin_session: str = Cookie(None)):
    """FastAPI dependency that blocks unauthenticated requests."""
    if not KEVIN_PASSWORD:
        raise HTTPException(status_code=503, detail="السيرفر غير مكوّن (KEVIN_PASSWORD مفقود)")
    if not verify_session_token(kevin_session or ""):
        raise HTTPException(status_code=401, detail="غير مسجل الدخول")
    return True

# ============================================
# INITIALIZATION
# ============================================
print(f"\n🚀 جاري تشغيل {ASSISTANT_NAME}...")

app = FastAPI(title=f"{ASSISTANT_NAME} - Arabic AI Assistant")

app.add_middleware(
    CORSMiddleware, allow_origins=["*"], allow_credentials=True,
    allow_methods=["*"], allow_headers=["*"],
)

print(f"📥 جاري تحميل Whisper ({WHISPER_MODEL})...")
whisper_model = whisper.load_model(WHISPER_MODEL)
print("✅ تم تحميل Whisper")

claude_client = None
gemini_model_instance = None

if ANTHROPIC_API_KEY:
    try:
        claude_client = Anthropic(api_key=ANTHROPIC_API_KEY)
        print("✅ Claude متصل")
    except Exception as e:
        print(f"⚠️ Claude error: {e}")

if GEMINI_API_KEY:
    try:
        genai.configure(api_key=GEMINI_API_KEY)
        gemini_model_instance = genai.GenerativeModel(GEMINI_MODEL, system_instruction=SYSTEM_PROMPT)
        print("✅ Gemini متصل")
    except Exception as e:
        print(f"⚠️ Gemini error: {e}")

if not claude_client and not gemini_model_instance:
    print("⚠️ تحذير: لم يتم تكوين أي مزود AI")

def load_history():
    if CONVERSATION_FILE.exists():
        try:
            return json.loads(CONVERSATION_FILE.read_text(encoding="utf-8"))
        except Exception:
            return []
    return []

def save_history(history):
    CONVERSATION_FILE.write_text(json.dumps(history, ensure_ascii=False, indent=2), encoding="utf-8")

conversation_history = load_history()

# ============================================
# AI ROUTING
# ============================================
MAX_TOOL_ITERATIONS = 4

async def _run_tool(name: str, tool_input: dict) -> str:
    tool = TOOLS_REGISTRY.get(name)
    if not tool:
        return f"خطأ: أداة غير معروفة '{name}'"
    try:
        result = await tool["handler"](**(tool_input or {}))
        return str(result) if result is not None else ""
    except TypeError as e:
        return f"خطأ في باراميترات الأداة: {e}"
    except Exception as e:
        return f"خطأ في تنفيذ الأداة: {type(e).__name__}: {str(e)[:200]}"

async def chat_with_claude(message, history):
    if not claude_client:
        raise HTTPException(status_code=503, detail="Claude غير متاح")
    messages = [
        {"role": h["role"], "content": h["content"]}
        for h in history if h["role"] in ["user", "assistant"] and isinstance(h.get("content"), str)
    ]
    messages.append({"role": "user", "content": message})
    tools = list_claude_tools()

    sys_prompt = build_system_prompt()
    for _ in range(MAX_TOOL_ITERATIONS):
        response = await asyncio.to_thread(
            claude_client.messages.create,
            model=CLAUDE_MODEL, max_tokens=2048,
            system=sys_prompt, messages=messages, tools=tools,
        )
        if response.stop_reason == "tool_use":
            tool_uses = [b for b in response.content if getattr(b, "type", None) == "tool_use"]
            tool_results = []
            for tu in tool_uses:
                result = await _run_tool(tu.name, dict(tu.input or {}))
                tool_results.append({
                    "type": "tool_result",
                    "tool_use_id": tu.id,
                    "content": result,
                })
            messages.append({"role": "assistant", "content": [b.model_dump() for b in response.content]})
            messages.append({"role": "user", "content": tool_results})
            continue

        # Final text answer
        parts = [b.text for b in response.content if getattr(b, "type", None) == "text"]
        return "\n".join(p for p in parts if p) or "[رد Claude فاضي]"

    return "[Claude تجاوز عدد محاولات استدعاء الأدوات]"

# Arabic command/filler words that should be stripped before sending to DDG.
_SEARCH_NOISE_AR = [
    "اعطيني", "اعطنى", "ادينى", "ادّيني", "هاتلي", "هات لي", "وريني", "قولي",
    "ايه", "إيه", "كل", "بسرعة", "بسرعه", "كده", "كدا", "كدة",
    "ابحث عن", "ابحثلي", "ابحث لي", "ابحث في النت", "ابحث",
    "في النت", "على النت", "على الإنترنت", "على الانترنت",
    "آخر", "اخر", "أحدث", "احدث", "آخر الـ", "أخر",
    "اللي", "اللى", "النهارده", "النهاردة",
    "لو سمحت", "من فضلك", "ممكن", "ياريت", "يا ريت",
    "في نقاط", "نقاط", "مختصرة", "مختصره", "بإيجاز", "ملخص",
]
_SEARCH_NOISE_EN = ["please", "can you", "give me", "show me", "tell me about", "what are"]

def clean_search_query(message: str) -> str:
    """Strip command/filler words so DDG gets the actual topic."""
    q = message.strip()
    low = q.lower()
    for w in _SEARCH_NOISE_EN:
        if w in low:
            i = low.index(w)
            q = (q[:i] + " " + q[i+len(w):]).strip()
            low = q.lower()
    for w in _SEARCH_NOISE_AR:
        q = q.replace(w, " ")
    # Collapse whitespace
    q = " ".join(q.split())
    # If we stripped everything, fall back to the original
    return q if len(q) >= 3 else message.strip()[:120]

async def chat_with_gemini(message, history):
    if not gemini_model_instance:
        raise HTTPException(status_code=503, detail="Gemini غير متاح")

    # Gemini path doesn't have native tool calling wired this round.
    # Pre-fetch tool results based on detected intent and inject into the prompt.
    # Also prepend persistent memory context.
    enriched = message
    mem_block = memory_summary_for_prompt()
    if mem_block:
        enriched = f"{mem_block}\n\n---\n\n{message}"

    # Time intent — cheap, run before LLM, inject result
    if detect_time_intent(message):
        t = await get_current_time()
        enriched = (
            f"{enriched}\n\n---\n"
            f"🕒 الوقت الحالي الفعلي (استخدم البيانات دي حصرياً ومتـ hallucinate-ش):\n{t}\n"
            f"ردّ بالمعلومات دي مباشرة."
        )
    # Weather intent — parse city names locally (no LLM call) and support multi-city
    elif detect_weather_intent(message):
        cities = extract_cities_from_message(message)
        if len(cities) > 5:
            cities = cities[:5]
        # Fetch weather for each city in parallel
        results = await asyncio.gather(*[get_weather(c) for c in cities], return_exceptions=True)
        joined = "\n\n".join(
            r if isinstance(r, str) else f"خطأ في {c}: {r}"
            for c, r in zip(cities, results)
        )
        enriched = (
            f"{enriched}\n\n---\n"
            f"🌤️ الطقس الفعلي (استخدم البيانات دي حصرياً):\n{joined}\n"
            f"ردّ بأسلوب طبيعي. لو فيه أكتر من مدينة، اعرضهم بالترتيب."
        )
    # Calc intent — extract math expression, run safely, inject
    elif detect_calc_intent(message):
        try:
            expr_req = (
                "Extract ONLY a math expression from this request. "
                "Output the expression as-is (use *, /, +, -, %, **, parentheses). "
                "Convert percentages like '15%' to '0.15*'. No explanation.\n\n"
                f"Request: {message}\n\nExpression:"
            )
            er = await asyncio.to_thread(gemini_model_instance.generate_content, expr_req)
            expression = (er.text or "").strip().splitlines()[0][:200]
        except Exception:
            expression = ""
        if expression:
            calc_result = await calculate(expression)
            enriched = (
                f"{enriched}\n\n---\n"
                f"🧮 نتيجة الحساب الدقيقة:\n{calc_result}\n"
                f"اعرض الناتج بأسلوب طبيعي."
            )
    elif detect_image_intent(message):
        # Build an English prompt via Gemini, then generate image
        try:
            img_prompt_req = (
                "Convert the user's request below into a detailed ENGLISH image prompt "
                "(15-30 words, comma-separated descriptors of subject, style, lighting, colors, mood). "
                "Output ONLY the prompt — no explanations.\n\n"
                f"User: {message}\n\nImage prompt:"
            )
            resp = await asyncio.to_thread(gemini_model_instance.generate_content, img_prompt_req)
            image_prompt = (resp.text or "").strip().strip('"\'').splitlines()[0][:400] or message[:200]
        except Exception as e:
            print(f"⚠️ Gemini image-prompt extraction failed: {e}")
            image_prompt = message[:200]
        print(f"🎨 Gemini extracted image prompt: '{image_prompt}'")
        image_md = await generate_image(image_prompt)
        enriched = (
            f"{message}\n\n"
            f"---\n"
            f"🎨 صورة تم توليدها بناءً على طلب المستخدم (اعرضها كما هي في ردك):\n"
            f"{image_md}\n\n"
            f"اكتب النص المرافق بالعربية (تهنئة، شرح، أو وصف حسب طلب المستخدم)."
        )
    else:
        should_search, kind = detect_search_intent(message)
        if should_search:
            search_query = clean_search_query(message)
            print(f"🔍 Cleaned search query: '{search_query}'")
            search_results = await web_search(search_query, kind=kind, max_results=5)
            # If search itself errored or returned nothing useful, don't pretend it worked
            looks_failed = (
                search_results.startswith("خطأ")
                or "مفيش نتائج" in search_results
                or len(search_results.strip()) < 30
            )
            if looks_failed:
                enriched = (
                    f"{message}\n\n"
                    f"---\n"
                    f"⚠️ ملاحظة داخلية: البحث في النت رجع فاضي للـ query '{search_query}'. "
                    f"رد على المستخدم بأمانة بأنك مش لاقي نتائج محدّثة، وقدّم اللي عندك من تدريبك "
                    f"مع توضيح إن المعلومة قد تكون قديمة. **متقولش 'مشكلة تقنية'** لأن مفيش مشكلة، "
                    f"النت بس مرجعش حاجة."
                )
            else:
                enriched = (
                    f"{message}\n\n"
                    f"---\n"
                    f"📡 نتائج بحث حديثة من النت — استخدم البيانات دي للإجابة "
                    f"واذكر المصادر باختصار. متجاوبش من تدريبك القديم — استخدم النتائج دي حصرياً:\n\n"
                    f"{search_results}"
                )

    gemini_history = []
    for h in history:
        content = h.get("content")
        if not isinstance(content, str):
            continue
        if h["role"] == "user":
            gemini_history.append({"role": "user", "parts": [content]})
        elif h["role"] == "assistant":
            gemini_history.append({"role": "model", "parts": [content]})
    try:
        chat = gemini_model_instance.start_chat(history=gemini_history)
        response = await asyncio.to_thread(chat.send_message, enriched)
        return response.text
    except Exception as e:
        err_text = str(e)
        is_rate_limit = ("429" in err_text or "quota" in err_text.lower()
                         or "rate" in err_text.lower())
        if is_rate_limit and claude_client:
            print(f"⚠️ Gemini rate-limited → fallback to Claude")
            try:
                # Use the enriched message (with injected tool results) for context
                return await chat_with_claude(enriched, history)
            except Exception as ce:
                ce_text = str(ce)
                if "429" in ce_text:
                    return ("⚠️ كلا الـ AI providers (Gemini و Claude) وصلوا للحد الأقصى. "
                            "استني دقيقة أو غيّر للـ Claude يدوياً من الزرار فوق.")
                raise
        raise

# ============================================
# ENDPOINTS
# ============================================
@app.get("/", response_class=HTMLResponse)
async def serve_ui(kevin_session: str = Cookie(None)):
    if not verify_session_token(kevin_session or ""):
        return RedirectResponse(url="/login", status_code=303)
    return HTMLResponse(HTML_UI)

@app.get("/login", response_class=HTMLResponse)
async def serve_login(kevin_session: str = Cookie(None)):
    if verify_session_token(kevin_session or ""):
        return RedirectResponse(url="/", status_code=303)
    return HTMLResponse(LOGIN_HTML)

@app.post("/api/login")
async def api_login(payload: dict, response: Response):
    if not KEVIN_PASSWORD:
        raise HTTPException(status_code=503, detail="السيرفر غير مكوّن (KEVIN_PASSWORD مفقود)")
    submitted = (payload.get("password") or "").strip()
    if not submitted:
        raise HTTPException(status_code=400, detail="ادخل كلمة المرور")
    if not hmac.compare_digest(submitted, KEVIN_PASSWORD):
        raise HTTPException(status_code=401, detail="كلمة مرور خاطئة")
    token = make_session_token()
    response.set_cookie(
        key=SESSION_COOKIE_NAME, value=token, max_age=SESSION_TTL_SECONDS,
        httponly=True, samesite="lax", secure=False, path="/",
    )
    return {"status": "ok", "message": "تم تسجيل الدخول"}

@app.post("/api/logout")
async def api_logout(response: Response):
    response.delete_cookie(SESSION_COOKIE_NAME, path="/")
    return {"status": "ok"}

@app.get("/api/status")
async def status(_auth: bool = Depends(require_auth)):
    mem = _load_memory()
    return {
        "assistant_name": ASSISTANT_NAME,
        "assistant_name_ar": ASSISTANT_NAME_AR,
        "version": ASSISTANT_VERSION,
        "claude_available": claude_client is not None,
        "gemini_available": gemini_model_instance is not None,
        "whisper_loaded": True,
        "voices": list(ARABIC_VOICES.keys()),
        "default_voice": DEFAULT_VOICE,
        "history_count": len(conversation_history),
        "knowledge_base_loaded": len(knowledge_base.get("files", [])),
        "tools": list(TOOLS_REGISTRY.keys()),
        "memory_facts": len(mem.get("facts", [])),
    }

@app.post("/api/transcribe")
async def transcribe_audio(audio: UploadFile = File(...), _auth: bool = Depends(require_auth)):
    """Convert Arabic audio to text using Whisper."""
    content = await audio.read()
    audio_size = len(content)

    if audio_size < 1000:
        print(f"⚠️ تسجيل صغير: {audio_size} bytes")
        raise HTTPException(status_code=400, detail=f"التسجيل صغير ({audio_size}b). اتكلم 2-3 ثواني")

    print(f"📥 استلام صوت: {audio_size} bytes ({audio.content_type})")

    # Determine extension from content type
    ext = ".webm"
    if audio.content_type and "mp4" in audio.content_type:
        ext = ".mp4"
    elif audio.filename and "." in audio.filename:
        ext = "." + audio.filename.split(".")[-1]

    with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as f:
        f.write(content)
        temp_path = f.name

    try:
        print(f"🎯 جاري التحويل بـ Whisper ({ext})... [temp: {temp_path}]")
        if not os.path.exists(temp_path):
            raise HTTPException(status_code=500, detail=f"الملف المؤقت غير موجود: {temp_path}")
        result = whisper_model.transcribe(
            temp_path,
            language="ar",
            fp16=False,
            no_speech_threshold=0.4,
            condition_on_previous_text=False,
            temperature=0.0,
        )
        text = result["text"].strip()
        print(f"✅ النتيجة: '{text[:100]}'")

        if not text or len(text) < 2:
            raise HTTPException(status_code=400, detail="مسمعتش كلام واضح. اتكلم بصوت أعلى")

        return {"text": text, "language": "ar", "audio_size": audio_size}
    except HTTPException:
        raise
    except FileNotFoundError as e:
        import traceback
        print(f"❌ FileNotFoundError: {e}")
        traceback.print_exc()
        msg = str(e).lower()
        if "ffmpeg" in msg or "winerror 2" in msg or e.filename == "ffmpeg" or e.filename == "ffmpeg.exe":
            raise HTTPException(
                status_code=500,
                detail="FFmpeg مش موجود. شغل: pip install imageio-ffmpeg ثم أعد تشغيل Kevin"
            )
        raise HTTPException(status_code=500, detail=f"ملف مفقود: {e}")
    except Exception as e:
        import traceback
        print(f"❌ خطأ في Whisper: {type(e).__name__}: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"خطأ تقني: {type(e).__name__}: {str(e)[:120]}")
    finally:
        try:
            os.unlink(temp_path)
        except Exception:
            pass

@app.post("/api/chat")
async def chat(payload: dict, _auth: bool = Depends(require_auth)):
    message = payload.get("message", "").strip()
    provider = payload.get("provider", "gemini")
    if not message:
        raise HTTPException(status_code=400, detail="رسالة فارغة")
    try:
        if provider == "claude":
            response_text = await chat_with_claude(message, conversation_history)
        elif provider == "gemini":
            response_text = await chat_with_gemini(message, conversation_history)
        else:
            raise HTTPException(status_code=400, detail="مزود غير معروف")
        timestamp = datetime.now().isoformat()
        conversation_history.append({"role": "user", "content": message, "timestamp": timestamp, "provider": provider})
        conversation_history.append({"role": "assistant", "content": response_text, "timestamp": datetime.now().isoformat(), "provider": provider})
        save_history(conversation_history)
        return {"response": response_text, "provider": provider, "timestamp": timestamp}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"خطأ: {str(e)}")

@app.post("/api/speak")
async def text_to_speech(payload: dict, _auth: bool = Depends(require_auth)):
    text = payload.get("text", "").strip()
    voice_key = payload.get("voice", DEFAULT_VOICE)
    if not text:
        raise HTTPException(status_code=400, detail="نص فارغ")
    voice = ARABIC_VOICES.get(voice_key, ARABIC_VOICES[DEFAULT_VOICE])
    with tempfile.NamedTemporaryFile(suffix=".mp3", delete=False) as f:
        temp_path = f.name
    try:
        communicate = edge_tts.Communicate(text, voice)
        await communicate.save(temp_path)
        return FileResponse(temp_path, media_type="audio/mpeg", filename="response.mp3")
    except Exception as e:
        try: os.unlink(temp_path)
        except Exception: pass
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/clear_history")
async def clear_history(_auth: bool = Depends(require_auth)):
    global conversation_history
    conversation_history = []
    save_history([])
    return {"status": "تم مسح المحادثة"}

@app.post("/api/reload_knowledge")
async def reload_knowledge(_auth: bool = Depends(require_auth)):
    """Reload the knowledge base from Excel files."""
    global knowledge_base
    # Delete cache
    if KNOWLEDGE_BASE_FILE.exists():
        KNOWLEDGE_BASE_FILE.unlink()
    knowledge_base = load_knowledge_base()
    return {"status": "تم تحديث قاعدة المعرفة", "files_loaded": len(knowledge_base.get("files", []))}

# ============================================
# STARTUP
# ============================================
def get_local_ip():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "127.0.0.1"

if __name__ == "__main__":
    local_ip = get_local_ip()
    print("\n" + "=" * 64)
    print(f"  🤖 {ASSISTANT_NAME_AR} ({ASSISTANT_NAME}) v{ASSISTANT_VERSION} — المساعد الذكي")
    print("=" * 64)
    kb_count = len(knowledge_base.get("files", []))
    mem_count = len(_load_memory().get("facts", []))
    print(f"  🔧 Tools المتاحة ({len(TOOLS_REGISTRY)}): {', '.join(TOOLS_REGISTRY.keys())}")
    print(f"  📚 Knowledge base: {kb_count} ملف Excel")
    print(f"  💾 Memory facts: {mem_count}")
    print(f"  🧠 LLMs: Claude {'✓' if claude_client else '✗'}  •  Gemini {'✓' if gemini_model_instance else '✗'}")
    print(f"  🎙️ Whisper STT: medium  •  🔊 Edge-TTS: {DEFAULT_VOICE}")
    print("-" * 64)
    print(f"  💻 لابتوب:  http://localhost:{PORT}")
    print(f"  📱 موبايل:  http://{local_ip}:{PORT}")
    print("=" * 64 + "\n")
    uvicorn.run(app, host=HOST, port=PORT, log_level="info")
